import phonenumbers
from phonenumbers import geocoder
import asyncio

import zipfile
import mimetypes
from phonenumbers import NumberParseException
asyncio.set_event_loop(asyncio.new_event_loop())


from telethon import TelegramClient, events, Button
import os, re, csv, json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from flask import Flask
import threading

# ================= FLASK KEEP ALIVE =================
app = Flask(__name__)

@app.route("/")
def home():
    return "SPIDY VCF BOT IS RUNNING 🚀", 200

def run_flask():
    port = int(os.environ.get("PORT", 5000))  # Render / Railway port
    app.run(host="0.0.0.0", port=port)

# ================= CONFIG =================
api_id = 34958210
api_hash = "6923cd2c34591c8e26b30ade39c7518b"        # PUT API HASH
bot_token = "8302113399:AAG0Pw1do6_kGrWTXGpZl20gPtoVY0dJeco"       # PUT BOT TOKEN
ADMIN_ID = 6860983540


FORCE_SUB_CHANNEL = "SPIDY_W_S"   # ✅ ONLY USERNAME, NO @, NO LINK
FORCE_SUB_FILE = "force_sub.json"
force_sub_enabled = False
# ================= INIT =================
client = TelegramClient("spidy_vcf_bot", api_id, api_hash).start(bot_token=bot_token)
# ================= ADMIN COMMANDS =================
@client.on(events.NewMessage(pattern="/admin"))
async def admin_panel(event):
    if event.sender_id != ADMIN_ID:
        return

    commands = """
╔════════════════════════════╗
║      🤖 BOT ADMIN PANEL        ║
╚════════════════════════════╝

🔐 ADMIN COMMANDS
━━━━━━━━━━━━━━━━━━━━

➤ /access USER_ID DAYS  
   ┗ 🎟️ Grant subscription access to a user

➤ /access on  
   ┗ 🆓 Enable FREE MODE (all users get access)

➤ /access off  
   ┗ 🔒 Disable FREE MODE (subscription system resumes)

➤ /admin  
   ┗ 🛠️ Display the admin command panel

➤ /info all  
   ┗ 📊 View complete subscription & user list

➤ /forcesub on  
   ┗ 📢 Enable force channel subscription

➤ /forcesub off  
   ┗ 🚫 Disable force channel subscription
   
➤ /broadcast MESSAGE
   ┗ 📢 Send message to all users
   
➤ /users
   ┗ 👥 Check total users before broadcast

➤ /report
┗ 📊 Export full user & subscription report

👤 USER COMMANDS
━━━━━━━━━━━━━━━━━━━━

➤ /start  
   ┗ 🚀 Start the bot and initialize services

➤ /subscription  
   ┗ 💳 Check your subscription status

➤ /done  
   ┗ 📂 Finish uploading files for processing
   
➤ /help  
   ┗ ℹ️ Step-by-step guide to use this bot

⚡ FEATURES
━━━━━━━━━━━━━━━━━━━━

✨ Advanced VCF Editor  
✨ Smart Automation Tools  
✨ Premium & VIP Options  

➡️ **All advanced features are available via inline buttons**

━━━━━━━━━━━━━━━━━━━━
🔥 Powered by @SPIDYWS
    """
    await event.reply(commands)
    
    
    
# ---------- STATES ----------
WAIT_ANALYZE_FILE = "WAIT_ANALYZE_FILE"
# ================= STORAGE =================

subscriptions = {}   # user_id -> expiry datetime
user_state = {}
user_files = {}
user_data = {}

SUB_FILE = "subscriptions.json"
FREE_MODE_FILE = "free_mode.json"  # Store free-mode state

GROUP_ACCESS_FILE = "group_access.json"
group_access = set()


USERS_FILE = "users.json"
all_users = set()

def load_users():
    global all_users
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, "r") as f:
                all_users = set(json.load(f))
        except:
            all_users = set()

def save_users():
    with open(USERS_FILE, "w") as f:
        json.dump(list(all_users), f)

load_users()


USER_JOIN_FILE = "user_join.json"
user_join_date = {}

def load_user_join():
    global user_join_date
    if os.path.exists(USER_JOIN_FILE):
        try:
            with open(USER_JOIN_FILE, "r") as f:
                user_join_date = {
                    int(k): datetime.fromisoformat(v)
                    for k, v in json.load(f).items()
                }
        except:
            user_join_date = {}

def save_user_join():
    with open(USER_JOIN_FILE, "w") as f:
        json.dump(
            {str(k): v.isoformat() for k, v in user_join_date.items()},
            f
        )

load_user_join()
# ================= SUBSCRIPTION PERSIST =================
def load_subscriptions():
    global subscriptions
    if os.path.exists(SUB_FILE):
        try:
            with open(SUB_FILE, "r") as f:
                data = json.load(f)
                subscriptions = {
                    int(uid): datetime.fromisoformat(exp)
                    for uid, exp in data.items()
                }
        except:
            subscriptions = {}

def save_subscriptions():
    with open(SUB_FILE, "w") as f:
        json.dump(
            {str(uid): exp.isoformat() for uid, exp in subscriptions.items()},
            f
        )

load_subscriptions()


def load_group_access():
    global group_access
    if os.path.exists(GROUP_ACCESS_FILE):
        try:
            with open(GROUP_ACCESS_FILE, "r") as f:
                group_access = set(json.load(f))
        except:
            group_access = set()

def save_group_access():
    with open(GROUP_ACCESS_FILE, "w") as f:
        json.dump(list(group_access), f)

load_group_access()

# ================= FREE MODE =================
free_mode = False

def load_free_mode():
    global free_mode
    if os.path.exists(FREE_MODE_FILE):
        try:
            with open(FREE_MODE_FILE, "r") as f:
                data = json.load(f)
                free_mode = data.get("free_mode", False)
        except:
            free_mode = False

def save_free_mode():
    with open(FREE_MODE_FILE, "w") as f:
        json.dump({"free_mode": free_mode}, f)

load_free_mode()


def load_force_sub():
    global force_sub_enabled
    if os.path.exists(FORCE_SUB_FILE):
        try:
            with open(FORCE_SUB_FILE, "r") as f:
                force_sub_enabled = json.load(f).get("enabled", False)
        except:
            force_sub_enabled = False

def save_force_sub():
    with open(FORCE_SUB_FILE, "w") as f:
        json.dump({"enabled": force_sub_enabled}, f)

load_force_sub()
# ================= HELPERS =================
def init_user(uid):
    user_state.setdefault(uid, "MENU")
    user_files.setdefault(uid, [])
    user_data.setdefault(uid, {})

def cleanup(uid):
    for f in user_files.get(uid, []):
        try:
            if os.path.exists(f):
                os.remove(f)
        except:
            pass
    user_files[uid] = []
    user_state[uid] = "MENU"
    user_data[uid] = {}

# ⬇️ TYPE IT HERE (NO BLANK LINE ABOVE def)
def group_allowed(event):
    if event.is_private:
        return True
    return event.chat_id in group_access
    
def is_subscribed(uid):
    if free_mode:  # <-- FREE MODE IGNORES SUBSCRIPTIONS
        return True
    if uid not in subscriptions:
        return False
    if datetime.now() > subscriptions[uid]:
        subscriptions.pop(uid, None)
        save_subscriptions()
        return False
    return True

def split_name_number(text):
    m = re.match(r"(.*?)(\d+)?$", text)
    base = m.group(1)
    start = int(m.group(2)) if m.group(2) else 1
    return base, start
    
def detect_country_and_format(number):
    try:
        if not number.startswith("+"):
            number = "+" + number

        parsed = phonenumbers.parse(number, None)
        formatted = phonenumbers.format_number(
            parsed, phonenumbers.PhoneNumberFormat.E164
        )

        country = geocoder.description_for_number(parsed, "en")
        return formatted, country if country else "Unknown"

    except:
        return number, "Invalid / Unknown"

async def notify_admin(msg):
    await client.send_message(ADMIN_ID, msg)

async def show_menu(chat):
    await client.send_message(
        chat,
        "👇 Choose option",
        buttons=[
            [Button.inline("🧑🏻‍🔧 EDIT VCF", b"edit")],
            [Button.inline("🔪 SPLIT VCF", b"split")],
            [Button.inline("🧪 ADVANCE VCF EDITOR", b"advance")]
        ]
    )


from telethon.tl.functions.channels import GetParticipantRequest

async def is_user_joined(event):
    if not force_sub_enabled:
        return True
    try:
        await client(
            GetParticipantRequest(
                FORCE_SUB_CHANNEL,
                event.sender_id
            )
        )
        return True
    except:
        return False
        
        
        
def analyze_any_file(path):
    ext = os.path.splitext(path)[1].lower()

    if ext == ".vcf":
        return analyze_vcf(path)
    elif ext == ".txt":
        return analyze_txt(path)
    elif ext == ".csv":
        return analyze_csv(path)
    elif ext == ".xlsx":
        return analyze_xlsx(path)
    elif ext == ".zip":
        return analyze_zip(path)
    else:
        size = round(os.path.getsize(path) / 1024, 2)
        return (
            "🔍 FILE ANALYSIS\n━━━━━━━━━━━━━━━━━━\n\n"
            f"📁 File Name : {os.path.basename(path)}\n"
            f"📦 File Size : {size} KB\n"
            "⚠️ Unsupported / binary file\n\n"
            "━━━━━━━━━━━━━━━━━━"
        )


def analyze_vcf(path):
    contacts = numbers = invalid = junk = 0
    valid_nums = []
    countries = {}

    with open(path, "r", errors="ignore") as f:
        card = []
        for line in f:
            line = line.strip()
            if line == "BEGIN:VCARD":
                card = []
            card.append(line)

            if line == "END:VCARD":
                contacts += 1
                name_ok = False

                for l in card:
                    if l.startswith(("FN:", "N:")):
                        val = l.split(":",1)[-1].strip().lower()
                        if not val or val.isdigit():
                            junk += 1
                        name_ok = True

                    if l.startswith("TEL"):
                        numbers += 1
                        num = l.split(":",1)[-1].strip()
                        try:
                            p = phonenumbers.parse(
                                num if num.startswith("+") else "+"+num, None
                            )
                            if not phonenumbers.is_valid_number(p):
                                raise NumberParseException(0, "Invalid")

                            fnum = phonenumbers.format_number(
                                p, phonenumbers.PhoneNumberFormat.E164
                            )
                            valid_nums.append(fnum)

                            c = geocoder.description_for_number(p, "en") or "Unknown"
                            countries[c] = countries.get(c, 0) + 1
                        except:
                            invalid += 1

                if not name_ok:
                    junk += 1

    unique = len(set(valid_nums))
    dup = len(valid_nums) - unique

    lines = [
        "🔍 VCF ANALYSIS REPORT",
        "━━━━━━━━━━━━━━━━━━\n",
        f"📁 File Name : {os.path.basename(path)}\n",
        f"📇 Total Contacts    : {contacts}",
        f"☎️ Total Numbers     : {numbers}",
        f"✅ Unique Numbers    : {unique}",
        f"🔁 Duplicate Numbers : {dup}\n",
        "🌍 COUNTRY BREAKDOWN"
    ]

    for c, n in sorted(countries.items(), key=lambda x: x[1], reverse=True):
        lines.append(f"• {c:<10} : {n}")

    lines += [
        "\n❌ INVALID NUMBERS",
        f"• {invalid} invalid / unparsable numbers found",
        "\n📛 NAME ISSUES",
        f"• {junk} contacts have empty or junk names",
        "\n━━━━━━━━━━━━━━━━━━",
        "ℹ️ No changes were made to your file"
    ]

    return "\n".join(lines)


def analyze_txt(path):
    with open(path, "r", errors="ignore") as f:
        lines = f.readlines()

    nums = [l.strip() for l in lines if l.strip().isdigit()]
    empty = sum(1 for l in lines if not l.strip())

    return (
        "🔍 TXT FILE ANALYSIS\n━━━━━━━━━━━━━━━━━━\n\n"
        f"📁 File Name : {os.path.basename(path)}\n\n"
        f"📄 Total Lines     : {len(lines)}\n"
        f"☎️ Number Lines    : {len(nums)}\n"
        f"🔁 Duplicate Nos   : {len(nums) - len(set(nums))}\n"
        f"📭 Empty Lines     : {empty}\n\n"
        "━━━━━━━━━━━━━━━━━━\nℹ️ File not modified"
    )


def analyze_csv(path):
    with open(path, newline='', encoding="utf-8", errors="ignore") as f:
        rows = list(csv.reader(f))

    header = rows[0] if rows else []
    nums = []
    for r in rows[1:]:
        for c in r:
            if c.strip().isdigit():
                nums.append(c.strip())

    return (
        "🔍 CSV FILE ANALYSIS\n━━━━━━━━━━━━━━━━━━\n\n"
        f"📁 File Name : {os.path.basename(path)}\n\n"
        f"📊 Total Rows     : {len(rows)}\n"
        f"📊 Total Columns  : {len(header)}\n"
        f"📑 Headers        : {', '.join(header)}\n"
        f"☎️ Numbers Found  : {len(nums)}\n"
        f"🔁 Duplicates     : {len(nums) - len(set(nums))}\n\n"
        "━━━━━━━━━━━━━━━━━━\nℹ️ File not modified"
    )


def analyze_xlsx(path):
    wb = load_workbook(path)
    nums = []

    lines = [
        "🔍 EXCEL FILE ANALYSIS",
        "━━━━━━━━━━━━━━━━━━\n",
        f"📁 File Name : {os.path.basename(path)}\n",
        f"📘 Total Sheets : {len(wb.sheetnames)}\n"
    ]

    for s in wb.sheetnames:
        sh = wb[s]
        lines.append(f"• {s} → {sh.max_row} rows")
        for row in sh.iter_rows(values_only=True):
            for c in row:
                if c and str(c).isdigit():
                    nums.append(str(c))

    lines += [
        "",
        f"☎️ Numbers Found  : {len(nums)}",
        f"🔁 Duplicates     : {len(nums) - len(set(nums))}",
        "\n━━━━━━━━━━━━━━━━━━",
        "ℹ️ File not modified"
    ]

    return "\n".join(lines)


def analyze_zip(path):
    with zipfile.ZipFile(path, 'r') as z:
        files = z.namelist()

    types = {}
    for f in files:
        e = os.path.splitext(f)[1].lower() or "no_ext"
        types[e] = types.get(e, 0) + 1

    lines = [
        "🔍 ZIP FILE ANALYSIS",
        "━━━━━━━━━━━━━━━━━━\n",
        f"📁 File Name : {os.path.basename(path)}\n",
        f"📦 Total Files : {len(files)}\n"
    ]

    for t, n in types.items():
        lines.append(f"• {t.upper()} : {n}")

    lines += [
        "\n━━━━━━━━━━━━━━━━━━",
        "ℹ️ ZIP not extracted"
    ]

    return "\n".join(lines)
# ================= START =================
@client.on(events.NewMessage(pattern="/start"))
async def start(event):
    uid = event.sender_id
    user = await event.get_sender()

    # =========================
    # 1️⃣ USER DATA SAVE (IMMEDIATE)
    # =========================
    all_users.add(uid)
    save_users()

    if uid not in user_join_date:
        user_join_date[uid] = datetime.now()
        save_user_join()

    # =========================
    # 2️⃣ SUBSCRIPTION STATUS CALCULATE
    # =========================
    if free_mode:
        sub_status = "🟢 FREE MODE"
    elif is_subscribed(uid):
        expiry = subscriptions.get(uid)
        if expiry:
            remaining = expiry - datetime.now()
            days = remaining.days
            hours = remaining.seconds // 3600
            sub_status = f"✅ SUBSCRIBED ({days}d {hours}h left)"
        else:
            sub_status = "✅ SUBSCRIBED"
    else:
        sub_status = "❌ NO SUBSCRIPTION"

    # =========================
    # 3️⃣ ADMIN KO SABSE PEHLE NOTIFY
    # =========================
    await notify_admin(
        "✨ **Bot Started** ✨\n\n"
        f"🙍 User: @{user.username if user.username else 'No username'}\n"
        f"🧾 ID: {uid}\n"
        f"📆 Date: {datetime.now().strftime('%d %b %Y')}\n"
        f"🕒 Time: {datetime.now().strftime('%I:%M %p')}\n\n"
        f"💼 Access Status: {sub_status}"
    )

    # =========================
    # 4️⃣ FORCE CHANNEL CHECK
    # =========================
    if force_sub_enabled:
        joined = await is_user_joined(event)
        if not joined:
            await event.reply(
                "⚠️ **Access Required**\n\n"
                "Joining our official channel is mandatory to use this service.\n\n"
                "After joining, tap **I Joined** below to continue.",
                buttons=[
                    [Button.url("🔔 Join Channel", f"https://t.me/{FORCE_SUB_CHANNEL}")],
                    [Button.inline("✅ I Joined", b"check_join")]
                ],
                link_preview=False
            )
            return

    # =========================
    # 5️⃣ BLOCK USER (AGAR SUBSCRIPTION NAHI HAI)
    # =========================
    if not free_mode and not is_subscribed(uid):
        await event.reply(
            "❌ **NO ACTIVE SUBSCRIPTION**\n\n"
            "💳 Buy premium access 👉 @SPIDYWS"
        )
        return

    # =========================
    # 6️⃣ SHOW MAIN MENU
    # =========================
    init_user(uid)
    await show_menu(event.chat_id)
# ================= USER HELP =================
@client.on(events.NewMessage(pattern="/help"))
async def user_help(event):
    text = (
        "🤖 **SPIDY VCF BOT – USER GUIDE**\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"

        "🪜 **HOW TO USE THIS BOT (STEP BY STEP)**\n\n"

        "➤ /start\n"
        "┗ 🚀 Start the bot & open main menu\n\n"

        "➤ Select feature using buttons\n"
        "┗ 👆 Choose EDIT / SPLIT / ADVANCE options\n\n"

        "➤ Upload VCF / TXT / CSV / XLSX / ZIP file\n"
        "┗ 📤 Send your file as per selected option\n\n"

        "➤ `/done`\n"
        "┗ ✅ Finish uploading & start processing\n\n"

        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🧑🏻‍🔧 **VCF FEATURES**\n\n"

        "➤ EDIT VCF\n"
        "┗ ✏️ Rename contacts & files automatically\n\n"

        "➤ SPLIT VCF\n"
        "┗ 🔪 Split big VCF into smaller files\n\n"

        "➤ MERGE VCF\n"
        "┗ 🔗 Merge multiple VCF files into one\n\n"

        "➤ MAKE VCF\n"
        "┗ 📇 Create VCF directly from numbers\n\n"

        "➤ ADD NUMBERS\n"
        "┗ ➕ Add new numbers to existing VCF\n\n"

        "➤ TXT → VCF\n"
        "┗ ♻️ Convert TXT numbers into VCF\n\n"

        "➤ CSV → VCF\n"
        "┗ ♻️ Convert CSV file into VCF\n\n"

        "➤ VCF → TXT\n"
        "┗ 📄 Extract numbers from VCF to TXT\n\n"

        "➤ XLSX → TXT\n"
        "┗ 📊 Convert Excel file to TXT numbers\n\n"

        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🔍 **VCF / FILE ANALYZER**\n\n"

        "➤ ANALYZE FILE\n"
        "┗ 🧠 Smart detailed analysis (No file editing)\n\n"

        "📊 **ANALYSIS REPORT INCLUDES**\n"
        "┗ 📁 File Name & Size\n"
        "┗ 📇 Total Contacts\n"
        "┗ ☎️ Total Numbers\n"
        "┗ ✅ Unique Numbers\n"
        "┗ 🔁 Duplicate Numbers\n"
        "┗ ❌ Invalid / Broken Numbers\n"
        "┗ 🌍 Country-wise Breakdown\n"
        "┗ 📛 Empty / Junk Contact Names\n\n"

        "📂 **SUPPORTED FILE TYPES**\n"
        "┗ VCF • TXT • CSV • XLSX • ZIP\n\n"

        "ℹ️ **NOTE**\n"
        "┗ Analysis is READ-ONLY\n"
        "┗ Your original file is NOT modified\n\n"

        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🧪 **ADVANCE VCF EDITOR**\n\n"

        "➤ ADMIN + NAVY VCF\n"
        "┗ 👮‍♂️⚓ Create combined ADMIN & NAVY contacts VCF\n"
        "   • Send ADMIN numbers\n"
        "   • Send NAVY numbers\n"
        "   • Auto country detect\n"
        "   • Fixed names: ADMIN01…, NAVY01…\n"
        "   • Single VCF + Summary report\n\n"

        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "💳 **SUBSCRIPTION COMMANDS**\n\n"

        "➤ /subscription\n"
        "┗ 💳 Check your active plan & remaining time\n\n"

        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🛒 **BUY PREMIUM ACCESS**\n"
        "┗ 💬 Contact : @SPIDYWS\n\n"

        "✨ Easy • Fast • Professional VCF Tools"
    )

    await event.reply(text)

# ================= SUBSCRIPTION STATUS =================
@client.on(events.NewMessage(pattern="/subscription"))
async def subscription_status(event):
    uid = event.sender_id

    if free_mode:
        await event.reply("✅ FREE MODE ACTIVE\nAll features unlocked.")
        return

    if uid not in subscriptions:
        await event.reply(
            "❌ NO ACTIVE SUBSCRIPTION\n\n"
            "💳 BUY NOW :- @SPIDYWS"
        )
        return

    expiry = subscriptions.get(uid)
    remaining = expiry - datetime.now()

    if remaining.total_seconds() <= 0:
        subscriptions.pop(uid, None)
        save_subscriptions()
        await event.reply(
            "❌ YOUR SUBSCRIPTION HAS EXPIRED\n\n"
            "💳 BUY NOW :- @SPIDYWS"
        )
        return

    days = remaining.days
    hours = remaining.seconds // 3600

    await event.reply(
        "✅ SUBSCRIPTION DETAILS\n\n"
        f"📅 Expiry Date: {expiry}\n"
        f"⏳ Remaining: {days} days {hours} hours"
    )

# ================= ADMIN GROUP ACCESS =================
@client.on(events.NewMessage(pattern="/accesson"))
async def access_on(event):
    if not event.is_group:
        await event.reply("❌ This command works only in groups.")
        return

    if event.sender_id != ADMIN_ID:
        return

    group_access.add(event.chat_id)
    save_group_access()

    await event.reply(
        "✅ GROUP ACCESS ENABLED\n\n"
        "All group members can now use the bot here."
    )


@client.on(events.NewMessage(pattern="/accessoff"))
async def access_off(event):
    if not event.is_group:
        await event.reply("❌ This command works only in groups.")
        return

    if event.sender_id != ADMIN_ID:
        return

    group_access.discard(event.chat_id)
    save_group_access()

    await event.reply(
        "🚫 GROUP ACCESS DISABLED\n\n"
        "Bot is now disabled for all group members."
    )
    
# ================= MAIN HANDLER =================
@client.on(events.NewMessage(func=lambda e: not e.raw_text or not e.raw_text.startswith("/")))
async def handler(event):
    uid = event.sender_id
    chat = event.chat_id

    joined = True      
    
    text = (event.raw_text or "").strip()
    state = user_state.get(uid)
      
          # 🔴 YAHIN PE LIKHNA HAI (VERY IMPORTANT)

    if not group_allowed(event):
        return

    # 🔒 FORCE SUB CHECK
    if force_sub_enabled and not is_subscribed(uid):
        joined = await is_user_joined(event)
        if not joined:
            await event.reply(
                "⚠️ **Access Required**\n\n"
                "Please join our official channel to continue.",
                buttons=[
                    [Button.url(
                        "🔔 Join Now",
                        f"https://t.me/{FORCE_SUB_CHANNEL.replace('@','')}"
                    )],
                    [Button.inline("✅ I Joined", b"check_join")]
                ]
            )
            return

    # 🔐 SUBSCRIPTION CHECK
    if not is_subscribed(uid):
        await event.reply(
            "❌ YOU DON'T HAVE AN ACTIVE SUBSCRIPTION.\n\n"
            "💳 BUY NOW :- @SPIDYWS"
        )
        return

    init_user(uid)
    state = user_state[uid]

    # 📤 FILE UPLOAD
    if event.file:
        path = f"{uid}_{event.file.name}"
        await event.download_media(path)
        user_files[uid].append(path)
        await event.reply("📥 File received\n➡️ Send /done")
        return

    # 📝 TEXT HANDLING (states continue...)
     
    # ---------- EDIT VCF ----------
    if state == "WAIT_VCF_BASE":
        user_data[uid]["vcf_raw_name"] = text.strip()  # 👈 EXACT user input

    # ❌ yahan kuch aur change nahi
        base, start = split_name_number(text)  # Case-2 ke liye as it is
        user_data[uid]["vcf_base"] = base
        user_data[uid]["vcf_start"] = start

        user_state[uid] = "WAIT_CONTACT_BASE"
        await event.reply("👤 Send base CONTACT name (Spidy01)")
        return

    if state == "WAIT_CONTACT_BASE":
        cname_base, cname_start = split_name_number(text)

        total_files = len(user_files[uid])
        counter = cname_start

        for i, src in enumerate(user_files[uid]):

        # ✅ ONLY CHANGE — Case-1
            if total_files == 1:
                out = f"{user_data[uid]['vcf_raw_name']}.vcf"
            else:
            # 🔒 Case-2 → EXACT OLD BEHAVIOUR
                vbase = user_data[uid]["vcf_base"]
                vstart = user_data[uid]["vcf_start"]
                out = f"{vbase}{vstart + i:02}.vcf"

            with open(src, "r", errors="ignore") as r:
                cards = r.read().split("END:VCARD")

            with open(out, "w") as w:
                for card in cards:
                    if "BEGIN:VCARD" not in card:
                        continue

                    name = f"{cname_base}{counter:02}"
                    counter += 1

                    w.write("BEGIN:VCARD\n")
                    for line in card.splitlines():
                        if line.startswith("FN:"):
                            w.write(f"FN:{name}\n")
                        elif line.startswith("N:"):
                            w.write(f"N:{name};;;;\n")
                        elif not line.startswith("BEGIN:VCARD"):
                            w.write(line + "\n")
                    w.write("END:VCARD\n")

            await client.send_file(event.chat_id, out)
            os.remove(out)

        cleanup(uid)
        await show_menu(event.chat_id)
        return

    # ---------- SPLIT ----------
    if state == "WAIT_SPLIT_COUNT":
        count = int(text)
        src = user_files[uid][0]

        with open(src, "r", errors="ignore") as r:
            cards = [c for c in r.read().split("END:VCARD") if "BEGIN:VCARD" in c]

        for i in range(0, len(cards), count):
            out = f"split_{i // count + 1}.vcf"
            with open(out, "w") as w:
                for c in cards[i:i + count]:
                    w.write(c.strip() + "\nEND:VCARD\n")
            await client.send_file(chat, out)
            os.remove(out)

        cleanup(uid)
        await show_menu(chat)
        return

    # ---------- MAKE VCF ----------
# ---------- MAKE VCF ----------
    if state == "WAIT_MAKE_NUMBERS":
        out = "numbers.vcf"
        detected_countries = set()

        with open(out, "w") as w:
            for i, n in enumerate(text.splitlines(), 1):
                n = n.strip()
                if not n:
                    continue

                formatted, country = detect_country_and_format(n)
                detected_countries.add(country)

                w.write(
                    "BEGIN:VCARD\n"
                    "VERSION:3.0\n"
                    f"N:Contact{i};;;;\n"
                    f"FN:Contact{i}\n"
                    f"TEL:{formatted}\n"
                    "END:VCARD\n"
                )

        await client.send_file(chat, out)
        os.remove(out)

        if detected_countries:
            await event.reply(
                "🌍 Detected Countries\n\n" +
                "\n".join(f"• {c}" for c in detected_countries)
            )

        cleanup(uid)
        await show_menu(chat)
        return

    # ---------- ADD NUMBERS ----------
    if state == "WAIT_ADDNUM_NUMBERS":
        src = user_files[uid][0]
        out = "added_numbers.vcf"

        with open(src, "r", errors="ignore") as r:
            data = r.read()

        with open(out, "w") as w:
            w.write(data)
            for i, n in enumerate(text.splitlines(), 1):
                if n.strip().isdigit():
                    w.write(
                        "BEGIN:VCARD\nVERSION:3.0\n"
                        f"N:New{i};;;;\nFN:New{i}\nTEL:{n}\nEND:VCARD\n"
                    )

        await client.send_file(chat, out)
        os.remove(out)
        cleanup(uid)
        await show_menu(chat)
        return

    # ---------- TXT TO VCF ----------
   # ---------- TXT TO VCF ----------
    if state == "WAIT_TXT_FILE":
        for src in user_files[uid]:
            base = os.path.splitext(os.path.basename(src))[0]
            out = f"{base}.vcf"
            counter = 1

            with open(out, "w") as w:
                with open(src, "r", errors="ignore") as r:
                    for line in r:
                        num = line.strip()
                        if num.isdigit():
                            w.write(
                                "BEGIN:VCARD\n"
                                "VERSION:3.0\n"
                                f"N:Contact{counter};;;;\n"
                                f"FN:Contact{counter}\n"
                                f"TEL:{num}\n"
                                "END:VCARD\n"
                            )
                            counter += 1

            await client.send_file(chat, out)
            os.remove(out)

        cleanup(uid)
        await show_menu(chat)
        return
    # ---------- CSV TO VCF ----------
    # ---------- CSV TO VCF ----------
    if state == "WAIT_CSV_FILE":
        for src in user_files[uid]:
            base = os.path.splitext(os.path.basename(src))[0]
            out = f"{base}.vcf"

            with open(out, "w") as w:
                with open(src, encoding="utf-8") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if len(row) >= 2 and row[1].strip().isdigit():
                            w.write(
                                "BEGIN:VCARD\n"
                                "VERSION:3.0\n"
                                f"N:{row[0]};;;;\n"
                                f"FN:{row[0]}\n"
                                f"TEL:{row[1]}\n"
                                "END:VCARD\n"
                            )

            await client.send_file(chat, out)
            os.remove(out)

        cleanup(uid)
        await show_menu(chat)
        return
    # ---------- VCF TO TXT ----------
   # ---------- VCF TO TXT ----------
    if state == "WAIT_VCF_TO_TXT":
        for src in user_files[uid]:
            base = os.path.splitext(os.path.basename(src))[0]
            out = f"{base}.txt"
            numbers = []

            with open(src, "r", errors="ignore") as r:
                for line in r:
                    if line.startswith("TEL:"):
                        numbers.append(line.replace("TEL:", "").strip())

            with open(out, "w") as w:
                w.write("\n".join(numbers))

            await client.send_file(chat, out)
            os.remove(out)

        cleanup(uid)
        await show_menu(chat)
        return


 # ---------- SPLIT (PART) ----------
    if state == "WAIT_PART_SPLIT_COUNT":
        try:
            parts = int(text)
            if parts < 2:
                await event.reply("❌ Parts must be 2 or more")
                return
        except:
            await event.reply("❌ Send a valid number (2,3,4...)")
            return

        src = user_files[uid][0]
  
        with open(src, "r", errors="ignore") as r:
            cards = [
                c.strip()
                for c in r.read().split("END:VCARD")
                if "BEGIN:VCARD" in c
            ]

        total = len(cards)
        per_part = total // parts
        extra = total % parts

        index = 0
        summary = []

        for i in range(parts):
            count = per_part + (1 if i < extra else 0)
            out = f"part_{i+1}.vcf"

            with open(out, "w") as w:
                for c in cards[index:index + count]:
                    w.write(c + "\nEND:VCARD\n")

            index += count
            await client.send_file(chat, out)
            summary.append(f"📂 Part {i+1}: {count} contacts")
            os.remove(out)

        await event.reply(
        "✅ VCF SPLIT COMPLETED\n\n"
        f"📇 Total contacts: {total}\n"
        f"🪓 Total parts: {parts}\n\n"
        + "\n".join(summary)
        )

        cleanup(uid)
        await show_menu(chat)
        return
        
        
        
    if state == "WAIT_ADMIN_NUMBERS":
        admin_nums = [n.strip() for n in text.splitlines()      if n.strip()]
        if not admin_nums:
            await event.reply("❌ Please send valid ADMIN numbers")
            return

        user_data[uid]["admin_numbers"] =         admin_nums
        user_state[uid] = "WAIT_NAVY_NUMBERS"

        await event.reply(
        "⚓ **Send NAVY phone numbers**\n"
        "➤ One number per line"
        )
        return
        
       
       
       
    if state == "WAIT_NAVY_NUMBERS":
        navy_nums = [n.strip() for n in text.splitlines() 
    if  n.strip()]
        if not navy_nums:
            await event.reply("❌ Please send valid NAVY numbers")
            return

        admin_nums =     user_data[uid]["admin_numbers"]

        out = "ADMIN_NAVY.vcf"

        admin_count = 0
        navy_count = 0
        detected_countries = set()   # 🌍

        with open(out, "w") as w:
        # 👮‍♂️ ADMIN CONTACTS
            for i, num in enumerate(admin_nums, 1):
                formatted, country =     detect_country_and_format(num)
                detected_countries.add(country)

                w.write(
                    "BEGIN:VCARD\n"
                    "VERSION:3.0\n"
                    f"N:ADMIN{i:02};;;;\n"
                    f"FN:ADMIN{i:02}\n"
                    f"TEL:{formatted}\n"
                    "END:VCARD\n"
                )
                admin_count += 1

        # ⚓ NAVY CONTACTS
            for i, num in enumerate(navy_nums, 1):
                formatted, country =     detect_country_and_format(num)
                detected_countries.add(country)

                w.write(
                    "BEGIN:VCARD\n"
                    "VERSION:3.0\n"
                    f"N:NAVY{i:02};;;;\n"
                    f"FN:NAVY{i:02}\n"
                    f"TEL:{formatted}\n"
                    "END:VCARD\n"
                )
                navy_count += 1

        await client.send_file(chat, out)
        os.remove(out)

    # 📊 SUMMARY MESSAGE (WITH COUNTRY INFO)
        await client.send_message(
            chat,
        "✅ **ADMIN + NAVY VCF CREATED SUCCESSFULLY**\n\n"
        "📁 File: `ADMIN_NAVY.vcf`\n"
        f"👮‍♂️ ADMIN Contacts : {admin_count}\n"
        f"⚓ NAVY Contacts   : {navy_count}\n"
        f"📊 Total Contacts  : {admin_count + navy_count}\n\n"
        "🌍 **Detected Countries**\n"
        + "\n".join(f"• {c}" for c in detected_countries if c)
    )

        cleanup(uid)
        await show_menu(chat)
        return
       
# ================= BUTTONS =================
@client.on(events.CallbackQuery)
async def buttons(event):
    uid = event.sender_id
    init_user(uid)

    data = event.data.decode()
    
    # ✅ FORCE SUB "I JOINED" BUTTON
    if data == "check_join":
        joined = await is_user_joined(event)

        if not joined:
            await event.answer(
                "❌ You have not joined the channel yet.",
                alert=True
            )
            return

        await event.edit(
            "✅ **Access Granted**\n\n"
            "Thank you for joining!\n"
            "You can now use the bot.\n\n"
            "➡️ Send /start to continue."
        )
        return

    if data == "edit":
        user_files[uid] = []
        user_state[uid] = "WAIT_EDIT_FILES"
        await event.reply("📤 Upload VCF files\n➡️ Send /done")

    elif data == "split":
        user_files[uid] = []
        user_state[uid] = "WAIT_SPLIT_FILES"
        await event.reply("📤 Upload ONE VCF file\n➡️ Send /done")

    elif data == "advance":
        await event.edit(
        "🧪 **ADVANCE VCF EDITOR**\n\nChoose an option 👇",
        buttons=[
            [Button.inline("🔗 MERGE VCF", b"merge")],
            [Button.inline("➕ ADD NUMBERS", b"addnum")],
            [Button.inline("📇 MAKE VCF", b"makevcf")],
            [Button.inline("🪓 SPLIT VCF (IN PARTS)", b"splitpart")],
            [Button.inline("👮‍♂️ ADMIN + ⚓ NAVY VCF",b"admin_navy")],
            [Button.inline("🔍 ANALYZE FILE", b"analyze_any")],
            [Button.inline("♻️ TXT TO VCF", b"txt2vcf")],
            [Button.inline("♻️ CSV TO VCF", b"csv2vcf")],
            [Button.inline("♻️ VCF TO TXT", b"vcftotxt")],
            [Button.inline("♻️ XLSX TO TXT", b"xlsx2txt")],
            [Button.inline("⬅️ BACK", b"back_main")]
        ]
    )
    elif data == "merge":
        user_files[uid] = []
        user_state[uid] = "WAIT_MERGE"
        await event.reply("📤 Upload VCF files\n➡️ Send /done")

    elif data == "addnum":
        user_files[uid] = []
        user_state[uid] = "WAIT_ADDNUM_FILE"
        await event.reply("📤 Upload ONE VCF file\n➡️ Send /done")

    elif data == "makevcf":
        user_state[uid] = "WAIT_MAKE_NUMBERS"
        await event.reply("📱 Send phone numbers")
        
    elif data == "splitpart":
        user_files[uid] = []
        user_state[uid] = "WAIT_PART_SPLIT_FILE"
        await event.reply("📤 Upload ONE VCF file\n➡️ Then send /done")
        
    elif data == "admin_navy":
        user_data[uid] = {}
        user_state[uid] = "WAIT_ADMIN_NUMBERS"
        await event.reply(
        "👮‍♂️ **Send ADMIN phone numbers**\n"
        "➤ One number per line\n"
        "➤ Example:\n"
        "9876543210\n9123456789"
        )
        
    elif data == "analyze_any":
        user_files[uid] = []
        user_state[uid] = WAIT_ANALYZE_FILE
        await event.reply(
        "📤 Upload ANY file (VCF / TXT / CSV / XLSX / ZIP)\n➡️ Then send /done"
        )

    elif data == "txt2vcf":
        user_files[uid] = []
        user_state[uid] = "WAIT_TXT_FILE"
        await event.reply("📄 Upload TXT file\n➡️ Send /done")

    elif data == "csv2vcf":
        user_files[uid] = []
        user_state[uid] = "WAIT_CSV_FILE"
        await event.reply("📊 Upload CSV file\n➡️ Send /done")

    elif data == "vcftotxt":
        user_files[uid] = []
        user_state[uid] = "WAIT_VCF_TO_TXT"
        await event.reply("📤 Upload VCF files\n➡️ Send /done")

    elif event.data == b"xlsx2txt":
        user_files[uid] = []
        user_state[uid] = "WAIT_XLSX_TO_TXT"
        await event.reply("📤 Upload XLSX file\n➡️ Then send /done")
    

        
    elif data == "back_main":
        user_state[uid] = "MENU"
        user_files[uid] = []
        user_data[uid] = {}

        await event.edit(
        "👇 Choose option",
        buttons=[
            [Button.inline("🧑🏻‍🔧 EDIT VCF", b"edit")],
            [Button.inline("🔪 SPLIT VCF", b"split")],
            [Button.inline("🧪 ADVANCE VCF EDITOR", b"advance")]
        ]
    )
# ================= DONE =================
@client.on(events.NewMessage(pattern="/done"))
async def done(event):
    uid = event.sender_id

    if user_state.get(uid) == "WAIT_EDIT_FILES":
        user_state[uid] = "WAIT_VCF_BASE"
        await event.reply("📝 Send BASE VCF file name")

    elif user_state.get(uid) == "WAIT_SPLIT_FILES":
        user_state[uid] = "WAIT_SPLIT_COUNT"
        await event.reply("🔢 Send contacts per file")

    elif user_state.get(uid) == "WAIT_MERGE":
        out = "merged.vcf"
        with open(out, "w") as w:
            for f in user_files[uid]:
                with open(f, "r", errors="ignore") as r:
                    w.write(r.read())
        await client.send_file(event.chat_id, out)
        os.remove(out)
        cleanup(uid)
        await show_menu(event.chat_id)

    elif user_state.get(uid) == "WAIT_ADDNUM_FILE":
        user_state[uid] = "WAIT_ADDNUM_NUMBERS"
        await event.reply("📱 Send numbers")

    elif user_state.get(uid) == "WAIT_VCF_TO_TXT":
        chat = event.chat_id

        if not user_files.get(uid):
            await event.reply("❌ No VCF files uploaded")
            return

        for src in user_files[uid]:
            base = os.path.splitext(os.path.basename(src))[0]
            out = f"{base}.txt"
            numbers = []

            with open(src, "r", errors="ignore") as r:
                for line in r:
                    if line.startswith("TEL"):
                        num = line.split(":", 1)[-1].strip()
                        if num:
                            numbers.append(num)

            if not numbers:
                await event.reply(f"⚠️ No numbers found in {base}.vcf")
                continue

            with open(out, "w") as w:
                w.write("\n".join(numbers))

            await client.send_file(chat, out)
            os.remove(out)

        cleanup(uid)
        await show_menu(chat)
     
    elif user_state.get(uid) == "WAIT_TXT_FILE":
        chat = event.chat_id

        if not user_files.get(uid):
            await event.reply("❌ No TXT files uploaded")
            return

        for src in user_files[uid]:
            base = os.path.splitext(os.path.basename(src))[0]
            out = f"{base}.vcf"
            counter = 1

            with open(out, "w") as w:
                with open(src, "r", errors="ignore") as r:
                    for line in r:
                        num = line.strip()
                        if num.isdigit():
                            w.write(
                                "BEGIN:VCARD\n"
                                "VERSION:3.0\n"
                                f"N:Contact{counter};;;;\n"
                                f"FN:Contact{counter}\n"
                                f"TEL:{num}\n"
                                "END:VCARD\n"
                            )
                            counter += 1

            await client.send_file(chat, out)
            os.remove(out)

        cleanup(uid)
        await show_menu(chat)
        
@client.on(events.NewMessage(pattern="/done"))
async def done(event):
    uid = event.sender_id
    chat = event.chat_id
    state = user_state.get(uid)

    if state is None:
        await event.reply("❌ No operation in progress")
        return

    # TXT → VCF
    if state == "WAIT_TXT_FILE":
        if not user_files.get(uid):
            await event.reply("❌ No TXT files uploaded")
            return
        # your existing TXT → VCF code here
        ...
        
        
@client.on(events.NewMessage(pattern="/done"))
async def done(event):
    uid = event.sender_id
    chat = event.chat_id
    state = user_state.get(uid)

    if not state:
        await event.reply("❌ No operation in progress")
        return

    # ---------- PART SPLIT ----------
    if state == "WAIT_PART_SPLIT_FILE":
        if not user_files.get(uid):
            await event.reply("❌ No VCF file uploaded")
            return

        user_state[uid] = "WAIT_PART_SPLIT_COUNT"
        await event.reply("🔢 Into how many parts do you want to split? (2,3,4...)")
        return

    # ---------- XLSX → TXT ----------
    elif state == "WAIT_XLSX_TO_TXT":
        if not user_files.get(uid):
            await event.reply("❌ No XLSX file uploaded")
            return

        from openpyxl import load_workbook

        for src in user_files[uid]:
            base = os.path.splitext(os.path.basename(src))[0]
            out = f"{base}.txt"

            wb = load_workbook(src)
            sheet = wb.active
            numbers = []

            for row in sheet.iter_rows(values_only=True):
                if len(row) > 1 and row[1]:
                    num = str(row[1]).strip()
                    if num.isdigit():
                        numbers.append(num)

            if not numbers:
                await event.reply("⚠️ No valid numbers found")
                continue

            with open(out, "w") as w:
                w.write("\n".join(numbers))

            await client.send_file(chat, out)
            os.remove(out)

        cleanup(uid)
        await show_menu(chat)
        return

@client.on(events.NewMessage(pattern="/done"))
async def done(event):
    uid = event.sender_id
    state = user_state.get(uid)

    # -------- ANALYZE FILE MODE --------
    if state == WAIT_ANALYZE_FILE:
        files = user_files.get(uid, [])

        if not files:
            await event.reply("❌ No files uploaded")
            return

        status_msg = await event.reply(
            f"🔍 Analyzing {len(files)} file(s)...\nPlease wait ⏳"
        )

        for i, path in enumerate(files, start=1):
            try:
                report = analyze_any_file(path)
                header = f"📂 FILE {i}/{len(files)}\n\n"
                await event.reply(f"```\n{header}{report}\n```")
            except Exception as e:
                await event.reply(
                    f"❌ Error analyzing `{os.path.basename(path)}`\n{e}"
                )

        # ✅ auto delete status message
        try:
            await status_msg.delete()
        except:
            pass

        cleanup(uid)
        await show_menu(event.chat_id)
        return
        
# ================= ADMIN =================
# ================= ADMIN COMMANDS =================
@client.on(events.NewMessage(pattern="/admin"))
async def admin_panel(event):
    if event.sender_id != ADMIN_ID:
        return

    commands = """
🤖 Admin Commands List
...
    """
    await event.reply(commands)

# ⬇️ PASTE /info all RIGHT HERE
@client.on(events.NewMessage(pattern="/info all"))
async def info_all(event):
    if event.sender_id != ADMIN_ID:
        return

    now = datetime.now()
    active_subs = {}
    removed_users = []  # 👈 expired users ka record

    # 🔄 Check & remove expired subscriptions
    for uid, expiry in list(subscriptions.items()):
        if expiry <= now:
            # expired user details save karo
            try:
                user = await client.get_entity(uid)
                uname = f"@{user.username}" if user.username else "No username"
            except:
                uname = "Unknown"

            removed_users.append({
                "uid": uid,
                "username": uname,
                "expired": expiry
            })
            continue

        active_subs[uid] = expiry

    # 🔁 Update subscriptions (only active)
    subscriptions.clear()
    subscriptions.update(active_subs)
    save_subscriptions()

    msg = "📋 **Subscription Report**\n\n"

    # ✅ ACTIVE USERS
    if subscriptions:
        msg += "🟢 **Active Subscriptions**\n\n"

        for uid, expiry in subscriptions.items():
            try:
                user = await client.get_entity(uid)
                uname = f"@{user.username}" if user.username else "No username"
            except:
                uname = "Unknown"

            remaining = expiry - now
            days = remaining.days
            hours = remaining.seconds // 3600

            msg += (
                f"👤 {uname}\n"
                f"🆔 {uid}\n"
                f"⏳ Remaining: {days}d {hours}h\n"
                f"──────────────\n"
            )
    else:
        msg += "🟢 **Active Subscriptions**\n"
        msg += "❌ No active subscriptions\n\n"

    # 🧹 REMOVED / EXPIRED USERS
    if removed_users:
        msg += "\n🧹 **Expired & Removed Users**\n\n"
        for u in removed_users:
            msg += (
                f"👤 {u['username']}\n"
                f"🆔 {u['uid']}\n"
                f"⌛ Expired on: {u['expired'].strftime('%d %b %Y %I:%M %p')}\n"
                f"──────────────\n"
            )
    else:
        msg += "\n🧹 **Expired & Removed Users**\n"
        msg += "✅ No expired users found\n"

    await event.reply(msg)


@client.on(events.NewMessage(pattern="/access"))
async def access(event):
    global free_mode
    if event.sender_id != ADMIN_ID:
        return

    try:
        parts = event.raw_text.split()
        if len(parts) == 2 and parts[1].lower() == "on":
            free_mode = True
            save_free_mode()
            await event.reply("✅ FREE MODE ENABLED\nAll users can access features without subscription.")
            return
        elif len(parts) == 2 and parts[1].lower() == "off":
            free_mode = False
            save_free_mode()
            await event.reply("✅ FREE MODE DISABLED\nSubscription system re-enabled.")
            return

        # Normal subscription grant
        uid = int(parts[1])
        days = int(parts[2])
        expiry = datetime.now() + timedelta(days=days)
        subscriptions[uid] = expiry
        save_subscriptions()

        await event.reply(
            f"✅ Access granted\nUser: {uid}\nDays: {days}\nExpires: {expiry}"
        )

        await client.send_message(
            uid,
            f"✅ SUBSCRIPTION ACTIVE\n⏳ Valid for {days} days\nSend /start"
        )
    except:
        await event.reply("❌ Usage: /access USER_ID DAYS OR /access on/off")
        
@client.on(events.NewMessage(pattern="/forcesub"))
async def force_sub_control(event):
    global force_sub_enabled
    if event.sender_id != ADMIN_ID:
        return

    parts = event.raw_text.split()
    if len(parts) != 2:
        await event.reply(
            "❌ Usage:\n"
            "/forcesub on\n"
            "/forcesub off"
        )
        return

    if parts[1].lower() == "on":
        force_sub_enabled = True
        save_force_sub()
        await event.reply("✅ FORCE SUBSCRIPTION ENABLED")
    elif parts[1].lower() == "off":
        force_sub_enabled = False
        save_force_sub()
        await event.reply("🚫 FORCE SUBSCRIPTION DISABLED")
    else:
        await event.reply("❌ Usage:\n/forcesub on | off")



@client.on(events.NewMessage(pattern="/broadcast"))
async def broadcast(event):
    if event.sender_id != ADMIN_ID:
        return

    msg = event.raw_text.replace("/broadcast", "", 1).strip()
    if not msg:
        await event.reply(
            "❌ **Broadcast Usage Error**\n\n"
            "`/broadcast Your message here`"
        )
        return

    total_users = len(all_users)
    if total_users == 0:
        await event.reply("❌ No users found to broadcast")
        return

    sent = 0
    failed = 0
    start_time = datetime.now()

    status_msg = await event.reply(
        "📢 **BROADCAST STARTED**\n\n"
        f"👥 Total Users: `{total_users}`\n"
        "📤 Sending messages...\n"
        "⏳ Please wait..."
    )

    for uid in list(all_users):
        try:
            await client.send_message(uid, msg)
            sent += 1
            await asyncio.sleep(0.07)  # safe delay
        except:
            failed += 1

    end_time = datetime.now()
    duration = (end_time - start_time).seconds

    await status_msg.edit(
        "✅ **BROADCAST COMPLETED**\n\n"
        f"👥 Total Users     : `{total_users}`\n"
        f"📤 Successfully Sent : `{sent}`\n"
        f"❌ Failed           : `{failed}`\n"
        f"⏱ Time Taken       : `{duration} sec`\n\n"
        "📌 Status: Finished Successfully"
    )
    
    
    
@client.on(events.NewMessage(pattern="/users"))
async def total_users(event):
    if event.sender_id != ADMIN_ID:
        return

    total_users = len(all_users)

    active_subs = 0
    expired_subs = 0

    now = datetime.now()

    for uid, exp in subscriptions.items():
        if exp > now:
            active_subs += 1
        else:
            expired_subs += 1

    free_mode_status = "🟢 ENABLED" if free_mode else "🔴 DISABLED"
    force_sub_status = "🟢 ENABLED" if force_sub_enabled else "🔴 DISABLED"

    await event.reply(
        "📊 **BOT USER STATISTICS DASHBOARD**\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"

        f"👥 **Total Registered Users**\n"
        f"➤ `{total_users}` users\n\n"

        f"💳 **Subscription Overview**\n"
        f"➤ 🟢 Active Subs : `{active_subs}`\n"
        f"➤ 🔴 Expired Subs : `{expired_subs}`\n\n"

        f"⚙️ **Bot System Status**\n"
        f"➤ 🆓 Free Mode : {free_mode_status}\n"
        f"➤ 🔔 Force Sub : {force_sub_status}\n\n"

        "📢 **Broadcast Info**\n"
        "➤ Use `/broadcast Your message`\n\n"

        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "🤖 Admin Control Panel"
    )
    
@client.on(events.NewMessage(pattern="/report"))
async def export_users_full(event):
    if event.sender_id != ADMIN_ID:
        return

    if not all_users:
        await event.reply("❌ No users found in database")
        return

    now = datetime.now()

    all_file = "all_users.txt"
    sub_file = "subscribed_users.txt"

    all_lines = []
    sub_lines = []

    # ---------- HEADER ----------
    header = [
        "📊 BOT USER DATABASE REPORT",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
        f"📅 Export Date : {now.strftime('%d %b %Y | %I:%M %p')}",
        f"👑 Admin ID    : {ADMIN_ID}",
        f"👥 Total Users : {len(all_users)}",
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
        ""
    ]

    all_lines.extend(header)
    sub_lines.extend(header)

    all_count = 0
    sub_count = 0

    # ---------- USER LOOP ----------
    for uid in sorted(all_users):
        try:
            user = await client.get_entity(uid)
            username = f"@{user.username}" if user.username else "N/A"
            fname = user.first_name or "-"
            lname = user.last_name or "-"
        except:
            username = "Unknown"
            fname = "-"
            lname = "-"

        join_dt = user_join_date.get(uid)
        joined_on = join_dt.strftime('%d %b %Y | %I:%M %p') if join_dt else "Unknown"

        # ---------- SUBSCRIPTION CHECK ----------
        if uid in subscriptions and subscriptions[uid] > now:
            exp = subscriptions[uid]
            remaining = exp - now
            account_type = "PREMIUM"
            sub_status = "ACTIVE"
            expiry_text = exp.strftime('%d %b %Y | %I:%M %p')
            remaining_text = f"{remaining.days} days"
            is_active = True
        else:
            account_type = "FREE"
            sub_status = "INACTIVE"
            expiry_text = "N/A"
            remaining_text = "N/A"
            is_active = False

        block = [
            "--------------------------------------------------",
            f"User No        : {all_count + 1}",
            f"User ID        : {uid}",
            f"Username       : {username}",
            f"First Name     : {fname}",
            f"Last Name      : {lname}",
            f"Account Type   : {account_type}",
            f"Subscription   : {sub_status}",
            f"Joined On      : {joined_on}",
            f"Expiry Date    : {expiry_text}",
            f"Time Remaining : {remaining_text}",
            ""
        ]

        all_lines.extend(block)
        all_count += 1

        if is_active:
            sub_block = block.copy()
            sub_block[1] = f"User No        : {sub_count + 1}"
            sub_lines.extend(sub_block)
            sub_count += 1

    # ---------- WRITE FILES ----------
    with open(all_file, "w", encoding="utf-8") as f:
        f.write("\n".join(all_lines))

    with open(sub_file, "w", encoding="utf-8") as f:
        f.write("\n".join(sub_lines))

    # ---------- STYLISH CAPTION ----------
    caption = (
        "📊 USER EXPORT COMPLETED SUCCESSFULLY\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"📅 Export Date : {now.strftime('%d %b %Y | %I:%M %p')}\n\n"
        "👥 USER STATISTICS\n"
        "━━━━━━━━━━━━━━━\n"
        f"• Total Users        : {all_count}\n"
        f"• Active Subscribers : {sub_count}\n\n"
        "📁 GENERATED FILES\n"
        "━━━━━━━━━━━━━━━\n"
        "• 📄 all_users.txt\n"
        "  └─ Contains ALL registered users\n\n"
        "• 📄 subscribed_users.txt\n"
        "  └─ Contains ONLY active subscribers\n\n"
        "⚡ Generated via Admin Control Panel"
    )

    await client.send_file(
        event.chat_id,
        [all_file, sub_file],
        caption=caption
    )

    # ---------- CLEANUP ----------
    os.remove(all_file)
    os.remove(sub_file)
# ================= RUN =================
print("SPIDY Bot running...")

# 🔥 START FLASK SERVER (KEEP ALIVE)
threading.Thread(target=run_flask, daemon=True).start()

# 🤖 START TELEGRAM BOT
client.run_until_disconnected()